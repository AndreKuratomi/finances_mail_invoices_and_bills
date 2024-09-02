import os
import time

from django.core.mail import EmailMessage, mail_admins
from django.template.loader import render_to_string

from openpyxl import load_workbook
from pathlib import Path
from rest_framework.views import APIView
from tqdm import tqdm

from management_before_django.table_managements.modules.openpyxl_module import colect_data_and_reset_due_date, status_update
from management_before_django.table_managements.modules.paths_module import paths_with_file_name, paths_with_many_file_names

not_found = True
count = 0
while not_found:
    print(count)
    try:
        print("oi")
        from model_to_email.models import TableName
        not_found = False
    except ImportError:
        count += 1
        time.sleep(1)
        continue
print("TableName:", TableName)

from robot_sharepoint.modules.robots.robot_to_download_attachments import download_attachments_from_sharepoint
from robot_sharepoint.modules.robot_utils.join_reports import join_reports

from utils.functions.deleting_elements import do_we_have_things_to_delete
from utils.variables.envs import user_email, password, nfe_email, sheet, sharepoint_measurements_url, download_directory, host_email
from utils.variables.paths import edited_tables_path, models_file_path, raw_tables_path, reports_path
from utils.variables.report_files import not_found_list, sent_list, not_found_title, sent_title


import ipdb

# Table to work with:
table_data = TableName.objects.all()
print("table_data:", table_data)

print("Views:", __name__)


class EmailAttachByTable(APIView):
    def post(self, root_dir: str):
        print(f"root_dir: {root_dir}")
        try:
            # Not found list report creation:
            with reports_path.joinpath(not_found_list).open("w") as file:
                file.write(not_found_title)

            for row in tqdm(table_data, "Each line, each search and email"):
                cnpj = row.cnpj
                contato = row.contatos
                nfe = row.numero
                nome_do_cliente = row.nome_do_cliente
                referencias = row.referencias
                status = row.status
                valor_liquido = row.valor_liquido
                vencimento = row.dt_vencto

                # Lógica para bug dt-vencimento = 31-12-1969:
                print("vencimento:", vencimento)
                if vencimento[6:8] != "20":
                    coluna_dt_vencimento: int = 11
                                        
                    (contatos, complete_file_path_to_raw, file_path_to_raw) = paths_with_many_file_names(raw_tables_path)
                    # Workbooks:
                    workbook_all_raw_data = load_workbook(data_only=True, filename=file_path_to_raw)
                    all_raw_data = workbook_all_raw_data[sheet]

                    # PLANILHA EDITADA:
                    # Paths:
                    (complete_file_path_to_edited, file_path_to_edited) = paths_with_file_name(edited_tables_path)
                    # Workbooks:
                    workbook_all_edited_data = load_workbook(data_only=True, filename=file_path_to_edited)
                    all_edited_data = workbook_all_edited_data[sheet]

                    colect_data_and_reset_due_date(all_raw_data, all_edited_data, coluna_dt_vencimento, complete_file_path_to_edited, workbook_all_edited_data)
                    # ipdb.set_trace()
                    
                    raise ValueError("ERROR!: Verificar coluna 'Dt_Vencimento'! Ano alterado para antes de 2000!")

                if status == "Não enviado":

                    row_data = {
                        # "competencia_por_ano": "competencia_por_ano",
                        # "contact": contato,
                        "contact": "andrekuratomi@gmail.com",
                        "cnpj": cnpj, 
                        "nfe": nfe, 
                        "nome_do_cliente": nome_do_cliente,
                        "referencias": referencias,
                        "valor_liquido": valor_liquido, 
                        "vencimento": vencimento, 
                    }
                    # ipdb.set_trace()
                    
                    # Extração de mês e ano:
                    refs = str(row_data['referencias'])

                    # PLACING TABLE TO WORK WITH WITH SELENIUM ROBOT:
                    download_attachments_from_sharepoint(
                        user_email,
                        password,
                        sharepoint_measurements_url,
                        download_directory,
                        cnpj,
                        nfe,
                        refs
                    )

                    attachments_path = "/robot_sharepoint/attachments/"
                    full_attachments_path = root_dir + attachments_path

                    # Extract info from attachments:
                    path = Path(full_attachments_path)
                    tables_path_content = list(path.iterdir())

                    competencia_por_ano = "02/01/2024"
                    tipo_de_servico = ""
                    table_template = "table_template_deposito.html"

                    # ipdb.set_trace()
                    # NOT FOUND CNPJ AND/OR NFE:
                    if len(tables_path_content) <= 1:
                        
                        # Fill element not found list:
                        not_found_elem = f"CNPJ: {cnpj} and/or NFE {nfe}. \n"
                        with reports_path.joinpath(not_found_list).open("a") as file:
                            file.write(not_found_elem)

                    else:
                        pode_enviar_email = True

                        for file in tables_path_content:
                            if file.is_file():
                                string_file = str(file)

                                if string_file.endswith('.pdf') or string_file.endswith('.xlsx'):
                                    prefix = full_attachments_path
                                    filtered = string_file[len(prefix):]

                                    # NFE 17757 FIXO Janeiro24.pdf
                                    if filtered.endswith('.pdf'):
                                        try:
                                            print("FILTERED:", filtered)

                                            if filtered.startswith("NFE"):

                                                # Extração de informação pelo título da NE:
                                                tipo_de_servico = ""
                                                for charac in filtered[10:]: # a partir de 'NFE <número_nfe> '
                                                    if charac == ' ':
                                                        break
                                                    else:
                                                        tipo_de_servico += charac
                                                print("TIPO_DE_SERVICO:", tipo_de_servico)

                                                competencia_por_ano = ""
                                                count = 0
                                                for charac in filtered:
                                                    if count == 3:
                                                        if charac == '.':
                                                            break
                                                        else:
                                                            competencia_por_ano += charac
                                                    if charac == ' ':
                                                        count += 1
                                                print("COMPETENCIA_POR_ANO:", competencia_por_ano)

                                            elif filtered.startswith("BOLETO"):
                                                table_template = "table_template_boleto.html"

                                            else:
                                                pode_enviar_email = False

                                        except Exception as e:
                                            print(f"ERRO! Arquivo pdf encontrado '{filtered}' não é nem NFE nem BOLETO e não pode ser considerado.")

                            else:
                                print("Error! Verify the file.")

                        print("competencia_por_ano:", competencia_por_ano)
                        print("nome_do_cliente_data:", row_data['nome_do_cliente'])
                        print("table_template:", table_template)
                        print("tipo_de_servico:", tipo_de_servico)

                        if pode_enviar_email:

                            # Insert table to mail body:
                            mail_content = render_to_string(
                                table_template, {
                                    'competencia_por_ano': competencia_por_ano, 
                                    'contact': row_data['contact'], 
                                    'nfe': row_data['nfe'], 
                                    'nome_do_cliente':  row_data['nome_do_cliente'],
                                    'tipo_de_servico': tipo_de_servico,
                                    'valor_liquido': row_data['valor_liquido'],
                                    'vencimento': row_data['vencimento']
                                }
                            )

                            time.sleep(2)  # wait for file to be created

                            email = EmailMessage(
                                "Nota Fiscal Eletrônica - J&C Faturamento - {a1}  {a2}  ( {a3} )  NF -  -  - {a4}"
                                .format(
                                    a1=tipo_de_servico, 
                                    a2=competencia_por_ano,
                                    a3=row_data['nome_do_cliente'], 
                                    a4=row_data['nfe']
                                ), # SUBJECT
                                mail_content, # BODY
                                "{}".format(host_email), # FROM
                                [row_data['contact']], # TO
                                [
                                    nfe_email, 
                                    "andrekuratomi@gmail.com"], # BCC
                            )
                            
                            # Reading HTML tags:
                            email.content_subtype = 'html'

                            # Attach files to email:
                            for file in tables_path_content:
                                print(file)
                                stringfy = str(file)
                                if stringfy.endswith('.pdf') or stringfy.endswith('.xlsx'):
                                    email.attach_file(file)

                            email.send()
                            print("Email successfully sent! Check inbox.")

                            # # FORMATAÇÃO DE DATA:

                            # dia = (datetime.now()).strftime("%d/%m/%Y")
                            # horas = (datetime.now()).strftime("%H:%M:%S")

                            # admin_email_message = """\
                            #     <html>
                            #         <head></head>
                            #         <body>
                            #             <p>Notificação: Foi enviado email para o usuário %s às %s em %s.</p>
                            #             <br>
                                        
                            #             <h3>J&C</h3>
                            #         </body>
                            #     </html>
                            # """ % (nome_do_cliente, horas, dia)

                            # mail_admins(
                            #     "Aviso de envio de email - Usuário {b1}".format(b1=nome_do_cliente), 
                            #     "",
                            #     fail_silently=False,
                            #     html_message=admin_email_message
                            # )

                            # Fill element sent list:
                            sent_elem = f"CNPJ: {cnpj} and/or NFE {nfe}. \n"
                            with reports_path.joinpath(sent_list).open("a") as file:
                                file.write(sent_elem)

                            # STATUS UPDATE:
                            status_update(edited_tables_path, row_data)

                else:
                    continue

            # Deletar conteúdo de raw_table depois de processo de envio ser finalizado:
            do_we_have_things_to_delete(raw_tables_path, '.xlsx')

            print("A APLICAÇÃO TERMINOU O PROCESSO DE ENVIO DE EMAILS COM SUCESSO!")

        except Exception as e:
            print(f"error:Something went wrong: {e} ! Contact the dev!")
