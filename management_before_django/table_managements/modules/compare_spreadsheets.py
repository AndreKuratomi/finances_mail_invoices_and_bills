import pandas as pd

from openpyxl import load_workbook

import ipdb


def compare_spreadsheets(path_to_raw: str, path_to_edited: str, full_path_to_edited: str, sheet: str) -> None:
    updated_workbook = load_workbook(data_only=True, filename=path_to_raw)
    old_workbook = load_workbook(data_only=True, filename=path_to_edited)

    worksheet = old_workbook.active

    updated_sheet = updated_workbook[sheet]
    old_sheet = old_workbook[sheet]

    updated_data = []
    old_data = []

    # Extract data from updated table
    for row in updated_sheet.iter_rows(min_row=2, values_only=True):
        updated_data.append(row)

    # Extract data from old table
    for row in old_sheet.iter_rows(min_row=2, values_only=True):
        old_data.append(row)
    print("previous_old_data1:", len(old_data))
    print("updated_data:", len(updated_data))
    new_lines = [row for row in updated_data if row[6] not in [old_row[6] for old_row in old_data]]

    if new_lines:
        print("LETS WORK!")
        old_data.extend(new_lines)
        print("new_old_data2:", len(old_data))
        
        for lines in new_lines:
            old_sheet.append(lines)
        
        old_workbook.save(full_path_to_edited)
        old_workbook.close()
    # ipdb.set_trace()
    print("DONE!")
