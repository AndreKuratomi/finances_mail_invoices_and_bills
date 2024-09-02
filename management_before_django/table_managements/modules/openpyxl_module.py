import pandas as pd

from management_before_django.table_managements.modules.paths_module import paths_with_file_name

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from pathlib import Path
from tqdm import tqdm
from utils.functions.path_length import do_we_have_spreadsheets

import ipdb


def adding_contacts_column(all_data: Worksheet, contacts_data: Worksheet, workbook_all_data: Workbook, complete_file_path_to_raw: str) -> None:
    """Openpyxl for adding a column 'CONTACTS' and updates the table downloaded in 'raw_table/'."""

    column_names = [col.value for col in all_data[1]]

    if "CONTACTS" not in column_names:
        new_column = all_data.max_column + 1
        
        all_data.insert_cols(new_column)
        all_data.cell(row=1, column=new_column).value = "CONTATOS"

        # Comparing by CNPJ:
        for index, all_data_row in enumerate(all_data.iter_rows(min_row=2, values_only=True), start=2):
            for contacts_row in contacts_data.iter_rows(min_row=2, values_only=True):
                contacts_cnpj = contacts_row[4].replace('.', '').replace('/', '').replace('-', '')
                if contacts_cnpj == all_data_row[4]:
                    all_data.cell(row=index, column=new_column).value = contacts_row[7]
                    break
  
        workbook_all_data.save(complete_file_path_to_raw)


def adding_references_column(all_data: Worksheet, workbook_all_data: Workbook, complete_file_path_to_raw: str) -> None:
    """
        Openpyxl for adding a column 'REFERENCES' collecting info of month and year from the column 'EMISSION DATE'.
        Updates the table downloaded in 'raw_table/'.
    """

    column_names = [col.value for col in all_data[1]]

    if "REFERENCES" not in column_names:
        new_column = all_data.max_column + 1

        all_data.insert_cols(new_column)
        all_data.cell(row=1, column=new_column).value = "REFERENCES"

        # coletar informação da coluna J (Data de emissão), editar e inserir na coluna nova.
        for index, all_data_row in enumerate(all_data.iter_rows(min_row=2, values_only=True), start=2):
            data_editada = str(all_data_row[9])[5:7] + '-' + str(all_data_row[9])[:4]
            all_data.cell(row=index, column=new_column).value = data_editada

        workbook_all_data.save(complete_file_path_to_raw)
        workbook_all_data.close()


def compare_spreadsheets(path_to_raw: str, path_to_edited: str, full_path_to_edited: str, sheet: str) -> None:
    """
        Compares the table just downloaded with the edited table (wit 'status', 'reference', etc).
        If the new one has new lines this function compares both and adds these new lines to the edited table.
    """

    updated_workbook = load_workbook(data_only=True, filename=path_to_raw)
    old_workbook = load_workbook(data_only=True, filename=path_to_edited)

    worksheet = old_workbook.active

    updated_sheet = updated_workbook[sheet]
    old_sheet = old_workbook[sheet]

    updated_data = []
    old_data = []

    # Extract data from updated table
    for row in updated_sheet.iter_rows(min_row=2, values_only=True):
        updated_data.append(row)

    # Extract data from old table
    for row in old_sheet.iter_rows(min_row=2, values_only=True):
        old_data.append(row)

    new_lines = [row for row in updated_data if row[6] not in [old_row[6] for old_row in old_data]]

    if new_lines:
        print("LETS WORK!")
        old_data.extend(new_lines)
        
        for lines in new_lines:
            old_sheet.append(lines)
        
        old_workbook.save(full_path_to_edited)
        old_workbook.close()

    print("DONE!")


def colect_data_and_reset_due_date(all_data: Worksheet, all_edited_data: Worksheet, column_number: int, complete_file_path_to_edited: str, workbook_all_edited_data: Workbook) -> None:
    """
        Collects cells from column 'due_date' from table in raw_table/, stores them and reinserts them in the same column in table edited_table/.
    """

    for raw_row in tqdm(range(2, all_data.max_row + 1), "Converting 'due date' back to original dates..."):
        raw_datetime_cell = all_data.cell(row=raw_row, column=column_number).value

        edited_datetime_cell = all_edited_data.cell(row=raw_row, column=column_number)
        edited_datetime_cell.value = raw_datetime_cell

    workbook_all_edited_data.save(complete_file_path_to_edited)
    workbook_all_edited_data.close()


def adding_status_column(all_data: Worksheet, edited_path: Path, file_path_to_raw: str, workbook_all_data: Workbook, sheet: str) -> None:
    """Openpyxl for adding the column 'STATUS' to the table in 'edited_table/'."""

    col_names = [col.value for col in all_data[1]]

    if "STATUS" not in col_names:
        new_column = all_data.max_column + 1

        all_data.insert_cols(new_column)
        all_data.cell(row=1, column=new_column).value = "STATUS"

        for cell in range(2, all_data.max_row + 1):
            all_data.cell(row=cell, column=new_column).value = "Not sent"

        table_in_edited_table_path = do_we_have_spreadsheets(edited_path, 1)

        # COMPARE HERE:
        if table_in_edited_table_path:
            print("Yes, we do.")
            (complete_file_path_to_edited, file_path_to_edited) = paths_with_file_name(edited_path)

            # Update table just downloaded with column status:
            workbook_all_data.save(file_path_to_raw)
            workbook_all_data.close()

            compare_spreadsheets(file_path_to_raw, file_path_to_edited, complete_file_path_to_edited, sheet)

        else:
            print("No, we don't.")

            # Saving edited table path:
            edited_file_path = str(edited_path.resolve() / 'edited_table.xlsx')

            workbook_all_data.save(edited_file_path)
            workbook_all_data.close()


def test_contacts(table_sheet: Worksheet, workbook: Workbook, complete_file_path_to_edited: str) -> None:
    """Openpyxl for changing contacts to test emails."""

    for cell in range(2, table_sheet.max_row + 1):
        if cell % 2 == 0:
            table_sheet.cell(row=cell, column=24).value = "test_1@company.com"
        else:
            table_sheet.cell(row=cell, column=24).value = "test_2@company.com"
    
    workbook.save(complete_file_path_to_edited)
    workbook.close()


def workbook_for_pandas(table_sheet: Worksheet) -> pd.DataFrame:
    """From the edited table above with Openpyxl to Pandas Dataframe."""

    rows = list()

    # First row for titles:
    headers = [cell.value for cell in table_sheet[1]]
    headers = [headers[3], headers[4], headers[6],  headers[10], headers[18], headers[23], headers[24], headers[25]]

    # Other rows for content:
    for row in tqdm(table_sheet.iter_rows(min_row=2), "Filtering rows by columns D, E, G, K, S, X, Y and Z..."):
        rows.append([cell.value for i, cell in enumerate(row) if i in [3, 4, 6, 10, 18, 23, 24, 25]]) #IMPROVE

    # PANDAS!
    df = pd.DataFrame(rows, columns=headers)

    # Adding column ID to make django work:
    if 'ID' not in df.columns:
        ID = range(1, df.shape[0]+1)
        df.insert(0, "id", ID)
        df.set_index('id')

    # Editing NFE column to hide first character 0:
    df['Number'] = df['Number'].apply(lambda x : x[1:])

    # Editing date column to show only date in brazilian format:
    df['Due Date'] = pd.to_datetime(df['Due Date'], errors='coerce')
    df['Due Date'] = df['Due Date'].dt.tz_localize('UTC').dt.tz_convert('America/Sao_Paulo')
    df['Due Date'] = (df['Due Date'] + pd.Timedelta(days=1)).dt.strftime('%d/%m/%Y')

    print(df)

    return df


def status_update(edited_path: Path, row_data: dict) -> None:
    """Receives the tables' path and the dictionary row_data, looks for the original line by the invoice value and if found updates the STATUS value to 'Sent'."""

    (complete_file_path_to_edited, file_path_to_edited) = paths_with_file_name(edited_path)

    # OPENPYXL:
    workbook = load_workbook(data_only=True, filename=file_path_to_edited)
    worksheet = workbook.active

    # Find column index for invoice ('Number'):
    column_index_for_nfe = None
    column_index_for_status = None

    for cell in worksheet[1]:
        if cell.value == 'Number':
            column_index_for_nfe = cell.column

        elif cell.value == "STATUS":
            column_index_for_status = cell.column

    if column_index_for_nfe is None or column_index_for_status is None:
        raise ValueError("Column 'Number' not found in the worksheet")
    else:
        try:
            for row_to_update in range(2, worksheet.max_row + 1): 
                
                # When invoice number began with 0:
                like_ancient_invoice = '0' + row_data['invoice']

                if worksheet.cell(row=row_to_update, column=column_index_for_nfe).value == str(like_ancient_invoice):
                    worksheet.cell(row=row_to_update, column=column_index_for_status).value = "Sent"
                    break
 
            workbook.save(complete_file_path_to_edited)
            workbook.close()

        except Exception as e:
            print(e)
            raise FileNotFoundError("No row with this invoice found!")
