from openpyxl import load_workbook

from pathlib import Path

from management_before_django.table_managements.modules.take_path_from_directory import paths_with_file_name

import ipdb


def status_update(edited_path: Path, row_data: dict) -> None:
    """Receives the tables' path and row_data as dict, searches for the original row by NFE value and if found updates status to 'Enviado'."""

    (complete_file_path_to_edited, file_path_to_edited) = paths_with_file_name(edited_path)

    # OPENPYXL TO ADD STATUS COLUMN:
    workbook = load_workbook(data_only=True, filename=file_path_to_edited)
    worksheet = workbook.active

    # Find column index for NFE ('Numero'):
    column_index_for_nfe = None
    column_index_for_status = None

    for cell in worksheet[1]:
        # print("cell.value:", cell.value)
        if cell.value == 'Numero':
            # ipdb.set_trace()
            column_index_for_nfe = cell.column
            # print("cell.column:", cell.column)
            # print(column_index)
        elif cell.value == "STATUS":
            column_index_for_status = cell.column

    if column_index_for_nfe is None or column_index_for_status is None:
        raise ValueError("Column 'Numero' not found in the worksheet")
    else:
        try:
            for row_to_update in range(2, worksheet.max_row + 1): 
                like_ancient_nfe = '0' + row_data['nfe']
                if worksheet.cell(row=row_to_update, column=column_index_for_nfe).value == str(like_ancient_nfe):
                    print("row_data['nfe']:", row_data['nfe'])
                    worksheet.cell(row=row_to_update, column=column_index_for_status).value = "Enviado"
                    # ipdb.set_trace()
                    print("worksheet.cell(row=row_to_update, column=column_index_for_status).value:", worksheet.cell(row=row_to_update, column=column_index_for_status).value)
                    break

            workbook.save(complete_file_path_to_edited)
            workbook.close()

        except Exception as e:
            print(e)
            raise FileNotFoundError("No row with this NFE found!")
