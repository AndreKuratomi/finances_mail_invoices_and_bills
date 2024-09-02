import pandas as pd
import time

from openpyxl import load_workbook
from pathlib import Path
from tqdm import tqdm

from management_before_django.table_managements.modules.openpyxl_module import adding_contacts_column, adding_references_column, adding_status_column, test_contacts, workbook_for_pandas
from management_before_django.table_managements.modules.paths_module import paths_with_file_name, paths_with_many_file_names

from utils.variables.envs import sheet, sheet_contacts

import ipdb


def filter_table_column(raw_path: Path, edited_path: Path, sheet: str) -> pd.DataFrame:
    """Module for working with all scripts that use openpyxl providing paths and workbooks."""

    # JUST DOWNLOADED TABLES:
    # Paths:
    (contatos, complete_file_path_to_raw, file_path_to_raw) = paths_with_many_file_names(raw_path)

    # Workbooks:
    workbook_contacts_data = load_workbook(data_only=True, filename=contatos)
    contacts_data = workbook_contacts_data[sheet_contacts]

    workbook_all_raw_data = load_workbook(data_only=True, filename=file_path_to_raw)
    all_raw_data = workbook_all_raw_data[sheet]

    # Openpyxl functions:
    adding_contacts_column(all_raw_data, contacts_data, workbook_all_raw_data, complete_file_path_to_raw)
    adding_references_column(all_raw_data, workbook_all_raw_data, complete_file_path_to_raw)
    adding_status_column(all_raw_data, edited_path, file_path_to_raw, workbook_all_raw_data, sheet)


    # EDITED TABLE:
    # Paths:
    (complete_file_path_to_edited, file_path_to_edited) = paths_with_file_name(edited_path)

    # Workbooks:
    workbook_all_edited_data = load_workbook(data_only=True, filename=file_path_to_edited)
    all_edited_data = workbook_all_edited_data[sheet]

    # Openpyxl functions:
    test_contacts(all_edited_data, workbook_all_edited_data, complete_file_path_to_edited)
    pandas_dataframe = workbook_for_pandas(all_edited_data)

    return pandas_dataframe
