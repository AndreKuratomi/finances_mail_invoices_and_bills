import datetime
import ipdb
import os

from openpyxl import load_workbook, Workbook
import pandas as pd

from pathlib import Path

from unidecode import unidecode

from tqdm import tqdm

from errors.custom_exceptions import TooManyFilesError

# # PATH TO RAW TABLE:
# tables_path = Path("./raw_table/")

# # PATH TO FILTERED TABLE:
# filtered_tables_path = Path("./filtered_table/")

# def get_cell_properties(cell):
#     properties = {}
#     properties['value'] = cell.value
#     properties['background_color'] = cell.fill.fgColor.rgb if cell.fill.patternType != 'none' else None
#     properties['data_type'] = cell.data_type
#     return properties

def filter_table_column(path: Path, sheet: str):
    # print(path)
    tables_path_content = list(path.iterdir())  

    tables = list()
    for elem in tables_path_content:
        stringfied_elem = str(elem)
        if stringfied_elem.endswith('.xls') or stringfied_elem.endswith('.xlsx') or stringfied_elem.endswith('.xlsm'):
            tables.append(stringfied_elem)
    if len(tables) == 0:
        raise FileNotFoundError("NO TABLE TO WORK WITH!")
    elif len(tables) > 1:
        raise TooManyFilesError
    
    for file in tables_path_content:
        if file.is_file():
            path_to_table = str(file)
            if path_to_table.endswith('.xls') or path_to_table.endswith('.xlsx') or path_to_table.endswith('.xlsm'):

                # OPENPYXL (to deal with colors):
                workbook = load_workbook(data_only=True, filename=path_to_table)
                table_sheet = workbook[sheet]

                # First row and all rows with yellow cells in colors list and changes saved:
                # new_sheet = workbook.create_sheet('filtered_sheet')
                # for_pandas = table_sheet.iter_cols(min_col=1, max_col=19)
                rows = list()
                headers = [cell.value for cell in table_sheet[2]]
                headers = [headers[3], headers[4], headers[6], headers[18]]
                # ipdb.set_trace()
                
                # # new_sheet.append([cell.value for cell in tqdm(table_sheet[1], "")])
                # counter = 0
                for row in tqdm(table_sheet.iter_rows(min_row=2), "Filtering rows by columns D, E and S..."):
                    # print(row)
                    # print(counter)
                    rows.append([cell.value for i, cell in enumerate(row) if i in [3, 4, 6, 18]]) #IMPROVE
                
                df = pd.DataFrame(rows, columns=headers)
                # print(df)
                
                # Drop first row:
                df.drop(0)

                # Adding column ID to make django work:
                if 'ID' not in df.columns:
                    ID = range(1, df.shape[0]+1)
                    df.insert(0, "id", ID)
                    df.set_index('id')

                # Editing NFE column to hide first character:
                df.loc[1:, 'Numero'] = df.loc[1:, 'Numero'].apply(lambda x : x[1:])

                return df

            else:
                raise Exception('Only .xls, .xlsx, or .xlsm files are supported.')
        else:
            raise Exception("Something went wrong... ")
