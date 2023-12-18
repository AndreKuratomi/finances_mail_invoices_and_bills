from openpyxl import load_workbook, Workbook
import pandas as pd

from pathlib import Path

from unidecode import unidecode

from tqdm import tqdm

import datetime
import ipdb
import os

# # PATH TO RAW TABLE:
# tables_path = Path("./raw_table/")

# # PATH TO FILTERED TABLE:
# filtered_tables_path = Path("./filtered_table/")

def get_cell_properties(cell):
    properties = {}
    properties['value'] = cell.value
    properties['background_color'] = cell.fill.fgColor.rgb if cell.fill.patternType != 'none' else None
    properties['data_type'] = cell.data_type
    return properties

# IN MEMORIAN LADY MARITACA (data_only=True):
def filter_table_by_yellow(path: Path, sheet: str):
    # print(path)
    tables_path_content = list(path.iterdir())  

    if len(tables_path_content) == 0:
        raise FileNotFoundError("NO TABLE TO WORK WITH!")
    elif len(tables_path_content) > 1:
        raise FileNotFoundError("HEY! ONLY ONE TABLE IS ALLOWED!")
    
    for file in tables_path_content:
        if file.is_file():
            path_to_table = str(file)
            if path_to_table.endswith('.xls') or path_to_table.endswith('.xlsx') or path_to_table.endswith('.xlsm'):

                # # Edit the character to find the table to work with not mattering the name:
                # specific_char = "/"
                # index = path_to_table.find(specific_char)
                # path_content = path.joinpath(path_to_table[index+1:])
                # # print(path_content)

                # OPENPYXL (to deal with colors):
                workbook = load_workbook(data_only=True, filename=path_to_table)
                table_sheet = workbook[sheet]

                # First row and all rows with yellow cells in colors list and changes saved:
                new_sheet = workbook.create_sheet('filtered_sheet')
                new_sheet.append([cell.value for cell in tqdm(table_sheet[1], "")])

                for row in tqdm(table_sheet.iter_rows(min_row=2), "Filtering rows by color..."):
                    filtered_row = [get_cell_properties(cell) for cell in row]
                    if filtered_row[3].get('background_color') == 'FFFFFF00':
                        # print(filtered_row)
                        new_sheet.append([cell.value for cell in row])
                
                # Preparing list for dataframe:
                rows = list()
                headers = [cell.value for cell in tqdm(new_sheet[1], "Inserting filtered tables' heardes into list before dataframe...")]

                for row in tqdm(new_sheet.iter_rows(min_row=2), "Inserting filtered tables into list before dataframe..."):
                    rows.append([cell.value for cell in row])
                
                df = pd.DataFrame(rows, columns=headers)
                
                # FIlter dataframe fromcolumn 0 to 13th:
                df = df.iloc[:, 0:13]
                # print(df)
                # ipdb.set_trace()

                if 'ID' not in df.columns:
                    ID = range(1, df.shape[0]+1)
                    df.insert(0, "id", ID)
                    df.set_index('id')

                # REMOVE CHARACTER ACCENTS FROM COLUMN TITLES:
                df.columns = [unidecode(col) for col in tqdm(df.columns, "Removing character accents from column titles...")]
                # print(df)

                # ORDER BY:
                df['PREVISAO DE CHEGADA'] = pd.to_datetime(df['PREVISAO DE CHEGADA'], errors='coerce')
                df = df.sort_values(by='PREVISAO DE CHEGADA', ascending=False)

                # print(df['METAL'])
                # print(df)
                return df

            else:
                raise Exception('Only .xls, .xlsx, or .xlsm files are supported.')
        else:
            raise Exception("Something went wrong... ")
